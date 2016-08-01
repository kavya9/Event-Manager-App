import click
import openpyxl
import MySQLdb
from warnings import filterwarnings

@click.group()
def mysqlCommands():
    filterwarnings('ignore', category=MySQLdb.Warning)# to remove warnings when using drop
    pass

#establishing connection

@mysqlCommands.command()
def importData():
    ''' import the data in the students.xlsx and marks.xlsx into the tables.'''
    db = MySQLdb.connect(host="localhost", user="root", passwd="kavya9",db="eventmanagement")
    cursor = db.cursor()
    studentsWb = openpyxl.load_workbook("student-Master-Data-2017.xlsx")

    sheets = studentsWb.get_sheet_names()
    for studentSheet in sheets:
        dept=studentSheet.encode('ascii','ignore')
        studentSheet = studentsWb.get_sheet_by_name(studentSheet)
        for rowNo in range(6, studentSheet.max_row + 1):
            studentName = studentSheet.cell(row=rowNo, column=3).value
            studentRollNumber = studentSheet.cell(row=rowNo, column=2).value
            if(dept=='MECH'): dept="mechanical"
            if (dept == 'Civil'): dept = "civil"
            if (dept == 'Chem'): dept = "chemical"

            studentDept=dept
            studentCGPA = (studentSheet.cell(row=rowNo, column=37).value)
            if studentCGPA is not None:
                studentCGPA = float(studentSheet.cell(row=rowNo, column=37).value)
            else:
                studentCGPA=0.0
            studentPassword="anits123*"
            attendance="0.0"

            #print studentCGPA,studentName,studentDept,studentPassword,studentRollNumber
            #print studentRollNumber
            cursor.execute('''insert into eventmanager_student(
                studentName,
                studentRollNumber,
                studentDept,
                studentCGPA,
                studentPassword,
                studentAttendance)
                values(%s,%s,%s,%s,%s,%s)''', (studentName, studentRollNumber, studentDept, float(studentCGPA),studentPassword,float(attendance)))
        db.commit()

if __name__=='__main__':
    mysqlCommands()




























































































































